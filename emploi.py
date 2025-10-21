import streamlit as st
import pdfplumber
import re
import pandas as pd
from io import BytesIO
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

st.title("🕕 Générateur d'emploi du temps par salle (PDF → Excel)")

cle = st.text_input("Donner la clé :")
if not cle:
    st.markdown("")
elif cle != "1234":
    st.error("La clé est incorrecte ❌")
else:

    seances = [
        ("08:15", "09:45"),
        ("09:55", "11:25"),
        ("11:35", "13:05"),
        ("13:10", "14:40"),
        ("14:50", "16:20"),
        ("16:30", "18:00")
    ]
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

    def extraire_cellule(cellule, jour, debut, fin, page_idx):
        return {
            "jour": jour,
            "debut": debut,
            "fin": fin,
            "contenu": cellule.strip().replace('\n', ' '),
            "page_idx": page_idx
        }

    def traiter_pdf(fichier_pdf):
        fichier_pdf.seek(0)
        resultats = []
        with pdfplumber.open(fichier_pdf) as pdf:
            nb_pages = len(pdf.pages)
            st.info(f"📄 PDF chargé avec succès : {nb_pages} pages.")
            for page_idx, page in enumerate(pdf.pages, start=1):
                st.write(f"🟢 Traitement de la page {page_idx}")
                tableaux = page.extract_tables()
                for tableau in tableaux:
                    if len(tableau) < 7:
                        continue
                    for i in range(min(len(seances), len(tableau) - 1)):
                        ligne = tableau[i + 1]
                        debut, fin = seances[i]
                        if len(ligne) < 7:
                            continue
                        for j in range(min(len(jours), len(ligne) - 1)):
                            cellule = ligne[j + 1]
                            if cellule and cellule.strip():
                                info = extraire_cellule(cellule, jours[j], debut, fin, page_idx)
                                resultats.append(info)
        return resultats

    def extraire_details(donnees):
        resultats, warnings = [], []
        for item in donnees:
            jour, debut, fin, contenu, page_idx = item["jour"], item["debut"], item["fin"], item["contenu"], item["page_idx"]
            lignes = [l.strip() for l in contenu.split('\n') if l.strip()]
            for ligne in lignes:
                pattern1 = r"(.+?)\s*-\s*([^-]+?)\s*-\s*([A-Za-z0-9.]+)\s*\(([^()]+(?:\(\*\))?)\)"
                match = re.match(pattern1, ligne)
                if match:
                    matiere, enseignant, salle, type_cours = (
                        match.group(1).strip(),
                        match.group(2).strip(),
                        match.group(3).strip(),
                        match.group(4).strip()
                    )
                else:
                    pattern2 = r"(.+?)\s+\(([^()]+)\)\s+([A-Za-z0-9.]+)\s+\(([^()]+(?:\(\*\))?)\)"
                    match = re.match(pattern2, ligne)
                    if match:
                        matiere, enseignant, salle, type_cours = (
                            match.group(1).strip(),
                            match.group(2).strip(),
                            match.group(3).strip(),
                            match.group(4).strip()
                        )
                    else:
                        warnings.append(f"⚠️ Non reconnu : {ligne}")
                        continue
                contenu_cellule = f"{matiere} - \n{enseignant} - \n{type_cours}"
                resultats.append({
                    "jour": jour,
                    "debut": debut,
                    "fin": fin,
                    "salle": salle,
                    "contenu": contenu_cellule
                })
        return resultats, warnings

    fichier_pdf = st.file_uploader("📂 Importer un fichier PDF", type=["pdf"])

    if fichier_pdf is not None:
        donnees_brutes = traiter_pdf(fichier_pdf)
        donnees_detaillees, warnings = extraire_details(donnees_brutes)

        salles_dict = defaultdict(list)
        for seance in donnees_detaillees:
            salles_dict[seance["salle"]].append(seance)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for salle, seances_salle in salles_dict.items():
                df = pd.DataFrame("", index=[f"{d}-{f}" for d, f in seances], columns=jours)
                for seance in seances_salle:
                    row_index = f"{seance['debut']}-{seance['fin']}"
                    col_index = seance["jour"]
                    df.at[row_index, col_index] = seance["contenu"]

                df.reset_index(inplace=True)
                df.rename(columns={"index": "Heure"}, inplace=True)
                df.to_excel(writer, index=False, sheet_name=salle[:31])

        output.seek(0)
        wb = load_workbook(output)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            ws.cell(row=1, column=1).value = f"Salle : {sheet}"
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)

            for i, col in enumerate(ws.columns, start=1):
                max_length = 0
                col_letter = get_column_letter(i)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 5

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    cell.font = Font(bold=True)
                    cell.border = thin_border

            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 30

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.download_button(
            label="📥 Télécharger l'emploi du temps",
            data=final_output.getvalue(),
            file_name="emploi_du_temps_salles.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if warnings:
            st.warning("Certaines lignes n'ont pas pu être reconnues.")
            st.text_area("⚠️ Détails :", value="\n".join(warnings), height=250)

    else:
        st.info("Merci de charger un fichier PDF pour commencer.")
